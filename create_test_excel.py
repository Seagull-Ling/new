# -*- coding: utf-8 -*-
"""
创建测试用Excel文件
包含30个中文姓名测试数据
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_test_excel():
    """创建测试用Excel文件"""
    
    # 30个中文姓名测试数据
    names = [
        '张伟', '王芳', '李娜', '刘洋', '陈杰',
        '杨敏', '赵磊', '黄婷', '周凯', '吴静',
        '徐涛', '孙丽', '胡勇', '朱琳', '高飞',
        '林燕', '何军', '郭婷', '马超', '罗丹',
        '梁爽', '宋倩', '郑辉', '谢娜', '韩磊',
        '唐静', '冯宇', '邓超', '曹颖', '彭博'
    ]
    
    # 部门数据
    departments = [
        '技术部', '市场部', '人事部', '财务部', '技术部',
        '市场部', '技术部', '人事部', '技术部', '市场部',
        '技术部', '财务部', '市场部', '人事部', '技术部',
        '市场部', '技术部', '财务部', '市场部', '人事部',
        '技术部', '市场部', '技术部', '财务部', '市场部',
        '人事部', '技术部', '市场部', '技术部', '财务部'
    ]
    
    # 职位数据
    positions = [
        '经理', '主管', '专员', '主管', '工程师',
        '专员', '主管', '经理', '工程师', '专员',
        '主管', '经理', '专员', '主管', '工程师',
        '专员', '主管', '工程师', '经理', '专员',
        '主管', '专员', '工程师', '主管', '经理',
        '专员', '主管', '工程师', '专员', '主管'
    ]
    
    # 创建DataFrame
    data = {
        '序号': list(range(1, 31)),
        '姓名': names,
        '部门': departments,
        '职位': positions
    }
    
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    output_path = r'd:\TREA\new\测试数据_30人.xlsx'
    
    # 使用openpyxl创建带格式的Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '人员名单'
    
    # 定义样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    headers = ['序号', '姓名', '部门', '职位']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 写入数据
    for row_idx, (idx, name, dept, pos) in enumerate(zip(data['序号'], data['姓名'], data['部门'], data['职位']), start=2):
        # 序号
        cell1 = ws.cell(row=row_idx, column=1, value=idx)
        cell1.alignment = center_align
        cell1.border = thin_border
        
        # 姓名
        cell2 = ws.cell(row=row_idx, column=2, value=name)
        cell2.alignment = center_align
        cell2.border = thin_border
        
        # 部门
        cell3 = ws.cell(row=row_idx, column=3, value=dept)
        cell3.alignment = center_align
        cell3.border = thin_border
        
        # 职位
        cell4 = ws.cell(row=row_idx, column=4, value=pos)
        cell4.alignment = center_align
        cell4.border = thin_border
    
    # 调整列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    wb.save(output_path)
    
    print(f"测试Excel文件已创建: {output_path}")
    print("\n文件内容预览:")
    print(df.to_string(index=False))
    
    return output_path


def create_test_excel_with_empty():
    """创建包含空值的测试Excel文件（用于测试空值处理）"""
    
    # 包含一些空值的测试数据
    names_with_empty = [
        '张伟', '', '李娜', None, '陈杰',
        '杨敏', '  ', '黄婷', None, '吴静',
        '徐涛', '', '胡勇', None, '高飞',
        '林燕', '  ', '郭婷', None, '罗丹',
        '梁爽', '', '郑辉', None, '韩磊',
        '唐静', '  ', '邓超', None, '彭博'
    ]
    
    # 创建DataFrame
    data = {
        '序号': list(range(1, 31)),
        '姓名': names_with_empty,
    }
    
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    output_path = r'd:\TREA\new\测试数据_含空值.xlsx'
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='人员名单', index=False)
    
    print(f"\n含空值的测试Excel文件已创建: {output_path}")
    
    return output_path


if __name__ == '__main__':
    print("=" * 60)
    print("创建智能排座系统测试数据")
    print("=" * 60)
    
    # 创建主测试文件
    create_test_excel()
    
    # 创建含空值的测试文件
    create_test_excel_with_empty()
    
    print("\n" + "=" * 60)
    print("测试数据创建完成！")
    print("=" * 60)
    print("\n使用说明:")
    print("1. 启动 Streamlit 应用: streamlit run app.py")
    print("2. 上传 '测试数据_30人.xlsx' 文件")
    print("3. 选择 '姓名' 列作为姓名顺序列")
    print("4. 配置每排人数并生成座位表")
